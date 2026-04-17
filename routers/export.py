import csv
import io
import datetime

from fastapi import APIRouter
from fastapi.responses import Response, StreamingResponse
from telethon import TelegramClient
from telethon.tl.types import MessageMediaPhoto, MessageMediaDocument
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from PIL import Image as PILImage

from schemas import ExportRequest
from services.telegram import get_session, ensure_connected, parse_date, parse_sender, fmt_date

router = APIRouter(tags=["export"])


def _is_image(msg) -> bool:
    return isinstance(msg.media, MessageMediaPhoto) or (
        isinstance(msg.media, MessageMediaDocument)
        and (getattr(msg.media.document, "mime_type", "") or "").startswith("image/")
    )


@router.post("/export")
async def export_posts(req: ExportRequest):
    sess = get_session(req.api_id)
    client: TelegramClient = sess["client"]
    await ensure_connected(client)

    date_from = parse_date(req.date_from)
    date_to = parse_date(req.date_to, end_of_day=True)

    rows = []
    for key, msg in sess.get("messages", {}).items():
        if not key.startswith(f"{req.chat_id}:"):
            continue
        if date_from and msg.date < date_from:
            continue
        if date_to and msg.date > date_to:
            continue
        rows.append((msg, parse_sender(msg.sender), _is_image(msg)))

    rows.sort(key=lambda x: x[0].date)

    if req.format == "csv":
        return _export_csv(rows, req.api_id, req.chat_id)
    return await _export_excel(client, rows, req.chat_id)


def _export_csv(rows: list, api_id: int, chat_id: str) -> Response:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["message_id", "date", "sender", "text", "image_url"])
    for msg, sender, is_image in rows:
        writer.writerow([
            msg.id,
            fmt_date(msg.date),
            sender,
            msg.text or msg.message or "",
            f"/media/{api_id}/{chat_id}/{msg.id}" if is_image else "",
        ])
    buf.seek(0)
    return Response(
        content=buf.getvalue().encode("utf-8-sig"),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=posts_{chat_id}.csv"},
    )


async def _export_excel(client: TelegramClient, rows: list, chat_id: str) -> StreamingResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "Posts"

    headers = ["Message ID", "Date", "Sender", "Text", "Image"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F6FEB")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["E"].width = 28

    IMG_W, IMG_H, ROW_H = 200, 200, 160

    for row_idx, (msg, sender, is_image) in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=msg.id)
        ws.cell(row=row_idx, column=2, value=fmt_date(msg.date))
        ws.cell(row=row_idx, column=3, value=sender)
        tc = ws.cell(row=row_idx, column=4, value=msg.text or msg.message or "")
        tc.alignment = Alignment(wrap_text=True, vertical="top")

        if is_image:
            img_data = await client.download_media(msg, file=bytes)
            if img_data:
                try:
                    pil = PILImage.open(io.BytesIO(img_data))
                    pil.thumbnail((IMG_W, IMG_H))
                    buf = io.BytesIO()
                    fmt = pil.format or "JPEG"
                    if fmt not in ("JPEG", "PNG", "GIF", "BMP"):
                        fmt = "JPEG"
                    pil.save(buf, format=fmt)
                    buf.seek(0)
                    xl_img = XLImage(buf)
                    xl_img.anchor = TwoCellAnchor(
                        _from=AnchorMarker(col=4, colOff=0, row=row_idx - 1, rowOff=0),
                        to=AnchorMarker(col=5, colOff=0, row=row_idx, rowOff=0),
                    )
                    ws.add_image(xl_img)
                except Exception:
                    ws.cell(row=row_idx, column=5, value="(load failed)")

        ws.row_dimensions[row_idx].height = ROW_H if is_image else 40

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=posts_{chat_id}.xlsx"},
    )
