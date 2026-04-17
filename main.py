from fastapi import FastAPI, HTTPException, Request
from fastapi.responses import Response, StreamingResponse
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel
from telethon import TelegramClient
from telethon.sessions import StringSession
from telethon.errors import SessionPasswordNeededError, PhoneCodeInvalidError
from telethon.tl.types import MessageMediaPhoto, MessageMediaDocument
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from PIL import Image as PILImage
import datetime, json, pathlib, io, csv, re

CONFIG_FILE = pathlib.Path("config.json")
templates = Jinja2Templates(directory="templates")

def parse_text(text: str) -> dict:
    """แยก roi_id และ ocr จาก text เช่น 'roi_id=roi03_1, ocr: 88'"""
    roi = re.search(r"roi_id\s*=\s*([^\s,]+)", text or "")
    ocr = re.search(r"ocr\s*:\s*(\d+(?:\.\d+)?)", text or "")
    return {
        "roi_id": roi.group(1) if roi else "",
        "ocr": ocr.group(1) if ocr else "",
    }
app = FastAPI()

# In-memory store per api_id session
_sessions: dict[int, dict] = {}


class ConfigModel(BaseModel):
    api_id: int
    api_hash: str
    phone: str
    chat_id: str | None = None
    date_from: str | None = None
    date_to: str | None = None
    sender_filter: str | None = None


@app.get("/config")
async def get_config():
    if CONFIG_FILE.exists():
        return json.loads(CONFIG_FILE.read_text())
    return {}


@app.post("/config")
async def save_config(cfg: ConfigModel):
    CONFIG_FILE.write_text(cfg.model_dump_json())
    return {"ok": True}


class SendCodeRequest(BaseModel):
    api_id: int
    api_hash: str
    phone: str


class VerifyCodeRequest(BaseModel):
    api_id: int
    api_hash: str
    phone: str
    code: str
    password: str | None = None  # 2FA password if enabled


class FetchRequest(BaseModel):
    api_id: int
    api_hash: str
    chat_id: str
    date_from: str | None = None
    date_to: str | None = None
    sender_filter: str | None = None


def _parse_date(s: str | None, end_of_day=False) -> datetime.datetime | None:
    if not s:
        return None
    d = datetime.datetime.strptime(s, "%Y-%m-%d")
    if end_of_day:
        d = d.replace(hour=23, minute=59, second=59)
    return d.replace(tzinfo=datetime.timezone.utc)


@app.post("/auth/send-code")
async def send_code(req: SendCodeRequest):
    client = TelegramClient(StringSession(), req.api_id, req.api_hash)
    await client.connect()
    try:
        result = await client.send_code_request(req.phone)
    except Exception as e:
        await client.disconnect()
        raise HTTPException(status_code=400, detail=str(e))

    _sessions[req.api_id] = {
        "client": client,
        "phone": req.phone,
        "phone_code_hash": result.phone_code_hash,
    }
    return {"ok": True, "message": "OTP sent"}


@app.post("/auth/verify")
async def verify_code(req: VerifyCodeRequest):
    sess = _sessions.get(req.api_id)
    if not sess:
        raise HTTPException(status_code=400, detail="Session not found, send code first")

    client: TelegramClient = sess["client"]
    try:
        await client.sign_in(
            phone=req.phone,
            code=req.code,
            phone_code_hash=sess["phone_code_hash"],
        )
    except SessionPasswordNeededError:
        if not req.password:
            raise HTTPException(status_code=400, detail="2FA password required")
        await client.sign_in(password=req.password)
    except PhoneCodeInvalidError:
        raise HTTPException(status_code=400, detail="OTP ไม่ถูกต้อง")
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

    sess["session_string"] = client.session.save()
    return {"ok": True, "message": "Login สำเร็จ"}


@app.post("/fetch")
async def fetch_messages(req: FetchRequest):
    sess = _sessions.get(req.api_id)
    if not sess or "session_string" not in sess:
        raise HTTPException(status_code=401, detail="ยังไม่ได้ login")

    client: TelegramClient = sess["client"]
    if not client.is_connected():
        await client.connect()

    date_from = _parse_date(req.date_from, end_of_day=False)
    date_to = _parse_date(req.date_to, end_of_day=True)

    try:
        entity = await client.get_entity(int(req.chat_id) if req.chat_id.lstrip("-").isdigit() else req.chat_id)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"ไม่พบ chat: {e}")

    posts = []
    # ดึงจาก date_to ย้อนหลัง หยุดเมื่อถึง date_from (เร็วกว่า reverse=True มาก)
    async for msg in client.iter_messages(entity, limit=None, offset_date=date_to):
        if date_to and msg.date > date_to:
            continue
        if date_from and msg.date < date_from:
            break

        sender_name = "unknown"
        if msg.sender:
            s = msg.sender
            if hasattr(s, "username") and s.username:
                sender_name = f"@{s.username}"
            elif hasattr(s, "first_name"):
                sender_name = s.first_name or "unknown"
            elif hasattr(s, "title"):
                sender_name = s.title

        # กรองตามชื่อผู้ส่งก่อนประมวลผลต่อ
        if req.sender_filter:
            keyword = req.sender_filter.lstrip("@").lower()
            if keyword not in sender_name.lstrip("@").lower():
                continue

        media_type = None
        is_image = False
        if msg.media:
            if isinstance(msg.media, MessageMediaPhoto):
                media_type = "photo"
                is_image = True
            elif isinstance(msg.media, MessageMediaDocument):
                doc = msg.media.document
                mime = getattr(doc, "mime_type", "") or ""
                if mime.startswith("image/"):
                    media_type = "image"
                    is_image = True
                else:
                    media_type = f"document ({mime or 'unknown'})"
            else:
                media_type = "media"

        # เก็บ message object ไว้สำหรับ /media endpoint
        if is_image:
            sess.setdefault("messages", {})[f"{req.chat_id}:{msg.id}"] = msg

        posts.append({
            "message_id": msg.id,
            "date": (msg.date + datetime.timedelta(hours=7)).strftime("%Y-%m-%d %H:%M:%S"),
            "sender": sender_name,
            "text": msg.text or msg.message or "(no text)",
            "media": media_type,
            "has_image": is_image,
        })

    posts.reverse()  # เรียงจากเก่าไปใหม่
    return {"count": len(posts), "posts": posts}


@app.get("/media/{api_id}/{chat_id}/{message_id}")
async def get_media(api_id: int, chat_id: str, message_id: int):
    sess = _sessions.get(api_id)
    if not sess:
        raise HTTPException(status_code=401, detail="ยังไม่ได้ login")

    key = f"{chat_id}:{message_id}"
    msg = sess.get("messages", {}).get(key)
    if not msg:
        raise HTTPException(status_code=404, detail="ไม่พบรูปภาพ")

    client: TelegramClient = sess["client"]
    if not client.is_connected():
        await client.connect()

    data = await client.download_media(msg, file=bytes)
    if not data:
        raise HTTPException(status_code=404, detail="ดาวน์โหลดไม่สำเร็จ")

    # ตรวจ mime type
    mime = "image/jpeg"
    if isinstance(msg.media, MessageMediaDocument):
        mime = getattr(msg.media.document, "mime_type", "image/jpeg") or "image/jpeg"

    return Response(content=data, media_type=mime)


class ExportRequest(BaseModel):
    api_id: int
    api_hash: str
    chat_id: str
    date_from: str | None = None
    date_to: str | None = None
    format: str = "excel"  # "excel" or "csv"



@app.post("/export")
async def export_posts(req: ExportRequest):
    sess = _sessions.get(req.api_id)
    if not sess or "session_string" not in sess:
        raise HTTPException(status_code=401, detail="ยังไม่ได้ login")

    client: TelegramClient = sess["client"]
    if not client.is_connected():
        await client.connect()

    # ดึง posts จาก messages cache ที่เก็บไว้ตอน fetch
    cached = sess.get("messages", {})
    date_from = _parse_date(req.date_from)
    date_to = _parse_date(req.date_to, end_of_day=True)

    rows = []
    for key, msg in cached.items():
        if not key.startswith(f"{req.chat_id}:"):
            continue
        if date_from and msg.date < date_from:
            continue
        if date_to and msg.date > date_to:
            continue

        sender = "unknown"
        if msg.sender:
            s = msg.sender
            if hasattr(s, "username") and s.username:
                sender = f"@{s.username}"
            elif hasattr(s, "first_name") and s.first_name:
                sender = s.first_name
            elif hasattr(s, "title"):
                sender = s.title

        is_image = isinstance(msg.media, MessageMediaPhoto) or (
            isinstance(msg.media, MessageMediaDocument)
            and (getattr(msg.media.document, "mime_type", "") or "").startswith("image/")
        )
        rows.append((msg, sender, is_image))

    rows.sort(key=lambda x: x[0].date)

    if req.format == "csv":
        return await _export_csv(rows, req.api_id, req.chat_id)
    else:
        return await _export_excel(client, rows, req.chat_id)


async def _export_csv(rows, api_id, chat_id):
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["message_id", "date", "sender", "text", "image_url"])
    for msg, sender, is_image in rows:
        img_url = f"/media/{api_id}/{chat_id}/{msg.id}" if is_image else ""
        writer.writerow([
            msg.id,
            (msg.date + datetime.timedelta(hours=7)).strftime("%Y-%m-%d %H:%M:%S"),
            sender,
            msg.text or msg.message or "",
            img_url,
        ])
    buf.seek(0)
    return Response(
        content=buf.getvalue().encode("utf-8-sig"),  # utf-8-sig สำหรับ Excel เปิดได้
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=posts_{chat_id}.csv"},
    )


async def _export_excel(client: TelegramClient, rows, chat_id):
    wb = Workbook()
    ws = wb.active
    ws.title = "Posts"

    # Header
    headers = ["Message ID", "Date", "Sender", "Text", "Image"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F6FEB")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["E"].width = 28

    IMG_W, IMG_H = 200, 200
    ROW_H_PX = 160

    for row_idx, (msg, sender, is_image) in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=msg.id)
        ws.cell(row=row_idx, column=2, value=(msg.date + datetime.timedelta(hours=7)).strftime("%Y-%m-%d %H:%M:%S"))
        ws.cell(row=row_idx, column=3, value=sender)
        text_cell = ws.cell(row=row_idx, column=4, value=msg.text or msg.message or "")
        text_cell.alignment = Alignment(wrap_text=True, vertical="top")

        if is_image:
            img_data = await client.download_media(msg, file=bytes)
            if img_data:
                try:
                    pil_img = PILImage.open(io.BytesIO(img_data))
                    pil_img.thumbnail((IMG_W, IMG_H))
                    img_buf = io.BytesIO()
                    fmt = pil_img.format or "JPEG"
                    if fmt not in ("JPEG", "PNG", "GIF", "BMP"):
                        fmt = "JPEG"
                    pil_img.save(img_buf, format=fmt)
                    img_buf.seek(0)
                    xl_img = XLImage(img_buf)
                    # col E = index 4 (0-based)
                    from_marker = AnchorMarker(col=4, colOff=0, row=row_idx - 1, rowOff=0)
                    to_marker   = AnchorMarker(col=5, colOff=0, row=row_idx,     rowOff=0)
                    anchor = TwoCellAnchor(_from=from_marker, to=to_marker)
                    xl_img.anchor = anchor
                    ws.add_image(xl_img)
                    ws.row_dimensions[row_idx].height = ROW_H_PX
                except Exception:
                    ws.cell(row=row_idx, column=5, value="(load failed)")
        ws.row_dimensions[row_idx].height = ROW_H_PX if is_image else 40

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=posts_{chat_id}.xlsx"},
    )


@app.get("/")
async def index(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})

